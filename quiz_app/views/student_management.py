"""
Student management views for admin users.

This module provides views for managing students including:
- Listing students
- Creating new students
- Editing existing students
- Deleting students
- Bulk import from CSV/Excel
- Bulk delete
"""
import csv
import io
import os
from datetime import date, datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Q
from django.core.paginator import Paginator
from django.db import transaction, IntegrityError
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from ..models import (
    Student,
    Form,
    LearningArea,
    SchoolInformation,
    User,
    Class,
    AcademicYear,
    StudentClass,
    TeacherSubjectAssignment,
    Teacher,
)


@login_required
@require_http_methods(["GET"])
def student_list_view(request):
    """
    Display list of all students with filtering and search capabilities.
    
    Supports:
    - Search by name, admission number, email
    - Filter by form, learning area, gender
    - Pagination
    - DataTables integration
    
    For teachers: Only shows students in classes they teach
    For admins/superadmins: Shows all students in their school
    """
    school = request.user.school
    students = Student.objects.all()
    
    # Handle teacher access - only show students in classes they teach
    if request.user.role == "teacher":
        if not hasattr(request.user, 'teacher_profile') or not request.user.teacher_profile:
            messages.error(request, "Teacher profile not found.")
            return redirect("quiz_app:dashboard")
        
        teacher = request.user.teacher_profile
        school = teacher.school
        
        # Get all active class assignments for this teacher
        teacher_assignments = TeacherSubjectAssignment.objects.filter(
            teacher=teacher,
            is_active=True
        )
        
        if school:
            teacher_assignments = teacher_assignments.filter(school=school)
        
        # Get the classes this teacher teaches
        teacher_classes = teacher_assignments.values_list('class_assigned', flat=True).distinct()
        
        if not teacher_classes:
            # Teacher has no class assignments, show empty list
            students = Student.objects.none()
        else:
            # Filter students to only those in classes the teacher teaches
            students = students.filter(
                studentclass__assigned_class__in=teacher_classes,
                studentclass__is_active=True
            )
            if school:
                students = students.filter(school=school)
            students = students.distinct()
    elif request.user.role in ["admin", "superadmin"]:
        # Admin and superadmin see all students in their school
        if school:
            students = students.filter(school=school)
    else:
        messages.error(request, "You don't have permission to access this page.")
        return redirect("quiz_app:dashboard")
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        students = students.filter(
            Q(full_name__icontains=search_query) |
            Q(admission_number__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(parent_contact__icontains=search_query)
        )
    
    # Filter by form (from current class assignment, not student's direct form)
    form_filter = request.GET.get('form', '')
    if form_filter:
        # Filter students who have an active assignment to a class with this form
        filter_kwargs = {
            'studentclass__assigned_class__form_id': form_filter,
            'studentclass__is_active': True,
        }
        if school:
            filter_kwargs['studentclass__school'] = school
        students = students.filter(**filter_kwargs).distinct()
    
    # Filter by learning area
    learning_area_filter = request.GET.get('learning_area', '')
    if learning_area_filter:
        students = students.filter(learning_area_id=learning_area_filter)
    
    # Filter by gender
    gender_filter = request.GET.get('gender', '')
    if gender_filter:
        students = students.filter(gender=gender_filter)
    
    # Filter by class
    class_filter = request.GET.get('class', '')
    if class_filter:
        # Filter students who have an active assignment to this class
        filter_kwargs = {
            'studentclass__assigned_class_id': class_filter,
            'studentclass__is_active': True,
        }
        if school:
            filter_kwargs['studentclass__school'] = school
        students = students.filter(**filter_kwargs).distinct()
    
    # Order by admission number
    students = students.order_by('-admission_date', 'full_name')
    
    # Prefetch current class assignments with class and form to optimize queries
    # This allows us to get the form from the assignment instead of student.form
    students = students.prefetch_related(
        'studentclass_set__assigned_class__form',
        'studentclass_set__assigned_class'
    ).select_related('form', 'learning_area')
    
    # Get filter options
    forms = Form.objects.all()
    if school:
        forms = forms.filter(school=school)
    
    learning_areas = LearningArea.objects.all()
    if school:
        learning_areas = learning_areas.filter(school=school)
    
    # Get classes for filter
    # For teachers, only show classes they teach
    if request.user.role == "teacher" and hasattr(request.user, 'teacher_profile') and request.user.teacher_profile:
        teacher = request.user.teacher_profile
        teacher_assignments = TeacherSubjectAssignment.objects.filter(
            teacher=teacher,
            is_active=True
        )
        if school:
            teacher_assignments = teacher_assignments.filter(school=school)
        teacher_classes = teacher_assignments.values_list('class_assigned', flat=True).distinct()
        classes = Class.objects.filter(id__in=teacher_classes)
        if school:
            classes = classes.filter(school=school)
        classes = classes.order_by('name')
    else:
        # Admin and superadmin see all classes
        classes = Class.objects.all()
        if school:
            classes = classes.filter(school=school)
        classes = classes.order_by('name')
    
    context = {
        'students': students,
        'forms': forms,
        'learning_areas': learning_areas,
        'classes': classes,
        'search_query': search_query,
        'form_filter': form_filter,
        'learning_area_filter': learning_area_filter,
        'gender_filter': gender_filter,
        'class_filter': class_filter,
    }
    
    # Use different template for teachers
    if request.user.role == "teacher":
        template = 'student/student_list_teacher.html'
    else:
        template = 'student/student_list.html'
    
    return render(request, template, context)


@login_required
@require_http_methods(["GET", "POST"])
def student_create_view(request):
    """
    Create a new student.
    
    GET: Returns form in modal
    POST: Creates student and returns JSON response
    """
    if request.user.role not in ["admin", "superadmin"]:
        if request.method == "GET":
            return JsonResponse({'error': 'Permission denied'}, status=403)
        messages.error(request, "You don't have permission to perform this action.")
        return redirect("quiz_app:student_list")
    
    school = request.user.school
    
    if request.method == "GET":
        # Return form for modal
        forms = Form.objects.all()
        learning_areas = LearningArea.objects.all()
        if school:
            forms = forms.filter(school=school)
            learning_areas = learning_areas.filter(school=school)
        
        html = render(request, 'student/partials/student_form.html', {
            'forms': forms,
            'learning_areas': learning_areas,
            'student': None,
        }).content.decode('utf-8')
        
        return JsonResponse({'html': html})
    
    # POST - Create student
    try:
        with transaction.atomic():
            # Get form data
            full_name = request.POST.get('full_name', '').strip()
            date_of_birth = request.POST.get('date_of_birth', '')
            gender = request.POST.get('gender', '')
            parent_contact = request.POST.get('parent_contact', '').strip()
            admission_date = request.POST.get('admission_date', '')
            email = request.POST.get('email', '').strip() or None
            form_id = request.POST.get('form', '') or None
            learning_area_id = request.POST.get('learning_area', '') or None
            profile_picture = request.FILES.get('profile_picture', None)
            
            # Validation
            if not all([full_name, date_of_birth, gender, parent_contact, admission_date]):
                return JsonResponse({
                    'success': False,
                    'error': 'Please fill in all required fields.'
                }, status=400)
            
            # Create student
            student = Student(
                full_name=full_name,
                date_of_birth=date_of_birth,
                gender=gender,
                parent_contact=parent_contact,
                admission_date=admission_date,
                email=email,
                school=school,
            )
            
            if form_id:
                try:
                    student.form = Form.objects.get(pk=form_id, school=school)
                except Form.DoesNotExist:
                    pass
            
            if learning_area_id:
                try:
                    student.learning_area = LearningArea.objects.get(pk=learning_area_id, school=school)
                except LearningArea.DoesNotExist:
                    pass
            
            if profile_picture:
                student.profile_picture = profile_picture
            
            student.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Student {student.full_name} created successfully.',
                'student_id': student.id,
                'admission_number': student.admission_number,
            })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error creating student: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["GET", "POST"])
def student_edit_view(request, student_id):
    """
    Edit an existing student.
    
    GET: Returns form in modal
    POST: Updates student and returns JSON response
    """
    if request.user.role not in ["admin", "superadmin"]:
        if request.method == "GET":
            return JsonResponse({'error': 'Permission denied'}, status=403)
        messages.error(request, "You don't have permission to perform this action.")
        return redirect("quiz_app:student_list")
    
    school = request.user.school
    
    # Superadmin can access any student, regular admin only their school
    if request.user.role == "superadmin":
        student = get_object_or_404(Student, pk=student_id)
    else:
        if not school:
            if request.method == "GET":
                return JsonResponse({'error': 'No school associated with your account'}, status=404)
            messages.error(request, "No school associated with your account.")
            return redirect("quiz_app:student_list")
        student = get_object_or_404(Student, pk=student_id, school=school)
    
    if request.method == "GET":
        # Return form for modal
        forms = Form.objects.all()
        learning_areas = LearningArea.objects.all()
        if school:
            forms = forms.filter(school=school)
            learning_areas = learning_areas.filter(school=school)
        
        html = render(request, 'student/partials/student_form.html', {
            'forms': forms,
            'learning_areas': learning_areas,
            'student': student,
        }).content.decode('utf-8')
        
        return JsonResponse({'html': html})
    
    # POST - Update student
    try:
        with transaction.atomic():
            # Get form data
            student.full_name = request.POST.get('full_name', '').strip()
            student.date_of_birth = request.POST.get('date_of_birth', '')
            student.gender = request.POST.get('gender', '')
            student.parent_contact = request.POST.get('parent_contact', '').strip()
            student.admission_date = request.POST.get('admission_date', '')
            student.email = request.POST.get('email', '').strip() or None
            
            form_id = request.POST.get('form', '') or None
            learning_area_id = request.POST.get('learning_area', '') or None
            
            if form_id:
                try:
                    student.form = Form.objects.get(pk=form_id, school=school)
                except Form.DoesNotExist:
                    student.form = None
            else:
                student.form = None
            
            if learning_area_id:
                try:
                    student.learning_area = LearningArea.objects.get(pk=learning_area_id, school=school)
                except LearningArea.DoesNotExist:
                    student.learning_area = None
            else:
                student.learning_area = None
            
            if 'profile_picture' in request.FILES:
                student.profile_picture = request.FILES['profile_picture']
            
            student.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Student {student.full_name} updated successfully.',
                'student_id': student.id,
            })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error updating student: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def student_delete_view(request, student_id):
    """
    Delete a student.
    """
    if request.user.role not in ["admin", "superadmin"]:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    school = request.user.school
    
    # Superadmin can access any student, regular admin only their school
    if request.user.role == "superadmin":
        student = get_object_or_404(Student, pk=student_id)
    else:
        if not school:
            return JsonResponse({'error': 'No school associated with your account'}, status=404)
        student = get_object_or_404(Student, pk=student_id, school=school)
    
    try:
        student_name = student.full_name
        student.delete()
        return JsonResponse({
            'success': True,
            'message': f'Student {student_name} deleted successfully.'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error deleting student: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["GET", "POST"])
def student_bulk_delete_view(request):
    """
    Bulk delete students.
    
    GET: Returns bulk delete modal
    POST: Deletes multiple students by IDs
    """
    if request.user.role not in ["admin", "superadmin"]:
        if request.method == "GET":
            return JsonResponse({'error': 'Permission denied'}, status=403)
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method == "GET":
        html = render(request, 'student/partials/bulk_delete_modal.html').content.decode('utf-8')
        return JsonResponse({'html': html})
    
    # POST - Bulk delete
    try:
        import json
        data = json.loads(request.body)
        student_ids = data.get('student_ids', [])
        
        if not student_ids:
            return JsonResponse({
                'success': False,
                'error': 'No students selected.'
            }, status=400)
        
        school = request.user.school
        students = Student.objects.filter(pk__in=student_ids)
        
        if school:
            students = students.filter(school=school)
        
        count = students.count()
        if count == 0:
            return JsonResponse({
                'success': False,
                'error': 'No valid students found to delete.'
            }, status=400)
        
        with transaction.atomic():
            students.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'{count} student(s) deleted successfully.'
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error deleting students: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["GET"])
def student_bulk_import_template_view(request):
    """
    Download a sample CSV template for student import.
    """
    if request.user.role not in ["admin", "superadmin"]:
        messages.error(request, "You don't have permission to perform this action.")
        return redirect("quiz_app:student_list")
    
    # Create a sample CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write headers - required: full_name, gender; optional: admission_number (auto-generated if not provided)
    headers = [
        'full_name',
        'gender',
        'admission_number',
        'date_of_birth',
        'parent_contact',
        'admission_date',
        'email',
    ]
    writer.writerow(headers)
    
    # Write sample data rows
    sample_rows = [
        [
            'John Doe',
            'M',
            'STU001',
            '2010-05-15',
            '0244123456',
            '2024-01-10',
            'john.doe@example.com',
        ],
        [
            'Jane Smith',
            'F',
            '',
            '2010-08-20',
            '0244987654',
            '2024-01-10',
            'jane.smith@example.com',
        ],
        [
            'Michael Johnson',
            'M',
            '',
            '2011-02-10',
            '0244111222',
            '2024-01-10',
            '',
        ]
    ]
    
    for row in sample_rows:
        writer.writerow(row)
    
    # Create HTTP response with CSV content
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="student_import_template.csv"'
    
    # Write CSV content to response
    output.seek(0)
    response.write(output.getvalue())
    output.close()
    
    return response


@login_required
@require_http_methods(["GET"])
def student_bulk_import_template_xlsx_view(request):
    """
    Download a sample XLSX template for student import.
    """
    if request.user.role not in ["admin", "superadmin"]:
        messages.error(request, "You don't have permission to perform this action.")
        return redirect("quiz_app:student_list")
    
    if not OPENPYXL_AVAILABLE:
        messages.error(request, "XLSX support is not available. Please install openpyxl.")
        return redirect("quiz_app:student_list")
    
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Import Template"
    
    # Define headers - required: full_name, gender; optional: admission_number (auto-generated if not provided)
    headers = [
        'full_name',
        'gender',
        'admission_number',
        'date_of_birth',
        'parent_contact',
        'admission_date',
        'email',
    ]
    
    # Style for header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers with styling
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Sample data rows - showing one with admission_number and one without (will be auto-generated)
    sample_rows = [
        [
            'John Doe',
            'M',
            'STU001',
            '2010-05-15',
            '0244123456',
            '2024-01-10',
            'john.doe@example.com',
        ],
        [
            'Jane Smith',
            'F',
            '',
            '2010-08-20',
            '0244987654',
            '2024-01-10',
            'jane.smith@example.com',
        ],
        [
            'Michael Johnson',
            'M',
            '',
            '2011-02-10',
            '0244111222',
            '2024-01-10',
            '',
        ]
    ]
    
    # Write sample data
    for row_num, row_data in enumerate(sample_rows, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Create HTTP response with XLSX content
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="student_import_template.xlsx"'
    
    # Save workbook to response
    wb.save(response)
    
    return response


def _read_file_headers(uploaded_file):
    """
    Helper function to read headers from CSV or XLSX file.
    Returns list of headers.
    """
    file_name = uploaded_file.name.lower()
    
    # Check if it's an Excel file
    if file_name.endswith(('.xlsx', '.xls')):
        if not OPENPYXL_AVAILABLE:
            raise ValueError("XLSX support is not available. Please install openpyxl.")
        
        from openpyxl import load_workbook
        
        # Read the Excel file
        uploaded_file.seek(0)  # Reset file pointer
        wb = load_workbook(uploaded_file, read_only=True, data_only=True)
        ws = wb.active
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value))
            else:
                break
        
        wb.close()
        uploaded_file.seek(0)  # Reset file pointer for later use
        return headers
    else:
        # Assume CSV
        uploaded_file.seek(0)
        file_data = uploaded_file.read().decode('utf-8')
        csv_reader = csv.DictReader(io.StringIO(file_data))
        headers = csv_reader.fieldnames or []
        uploaded_file.seek(0)  # Reset file pointer for later use
        return list(headers) if headers else []


@login_required
@require_http_methods(["POST"])
def student_bulk_import_preview_headers_view(request):
    """
    Preview CSV/XLSX headers for column mapping.
    """
    if request.user.role not in ["admin", "superadmin"]:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if 'csv_file' not in request.FILES:
        return JsonResponse({
            'success': False,
            'error': 'No file uploaded.'
        }, status=400)
    
    uploaded_file = request.FILES['csv_file']
    try:
        headers = _read_file_headers(uploaded_file)
        
        return JsonResponse({
            'success': True,
            'headers': headers
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error reading file: {str(e)}'
        }, status=400)


@login_required
@require_http_methods(["GET", "POST"])
def student_bulk_import_view(request):
    """
    Bulk import students from CSV or XLSX file.
    
    GET: Returns import form modal
    POST: Processes CSV/XLSX file and imports students
    """
    if request.user.role not in ["admin", "superadmin"]:
        if request.method == "GET":
            return JsonResponse({'error': 'Permission denied'}, status=403)
        messages.error(request, "You don't have permission to perform this action.")
        return redirect("quiz_app:student_list")
    
    if request.method == "GET":
        # Return import form for modal
        classes = Class.objects.all()
        school = request.user.school
        if school:
            classes = classes.filter(school=school)
        
        classes = classes.order_by('name')
        
        # Get current academic year to display in template
        current_academic_year = SchoolInformation.get_current_academic_year(school=school)
        
        html = render(request, 'student/partials/bulk_import_modal.html', {
            'classes': classes,
            'current_academic_year': current_academic_year,
        }).content.decode('utf-8')
        
        return JsonResponse({'html': html})
    
    # POST - Process import
    try:
        import json
        
        if 'csv_file' not in request.FILES:
            return JsonResponse({
                'success': False,
                'error': 'No file uploaded.'
            }, status=400)
        
        uploaded_file = request.FILES['csv_file']
        school = request.user.school
        
        # Determine file type and read accordingly
        file_name = uploaded_file.name.lower()
        
        if file_name.endswith(('.xlsx', '.xls')):
            # Handle Excel file
            if not OPENPYXL_AVAILABLE:
                return JsonResponse({
                    'success': False,
                    'error': 'XLSX support is not available. Please install openpyxl or use CSV format.'
                }, status=400)
            
            from openpyxl import load_workbook
            
            uploaded_file.seek(0)
            wb = load_workbook(uploaded_file, read_only=True, data_only=True)
            ws = wb.active
            
            # Get headers from first row
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value))
                else:
                    break
            
            # Convert Excel rows to dictionary format
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell for cell in row):  # Skip empty rows
                    row_dict = {}
                    for idx, header in enumerate(headers):
                        if idx < len(row):
                            row_dict[header] = str(row[idx]) if row[idx] is not None else ''
                        else:
                            row_dict[header] = ''
                    rows.append(row_dict)
            
            wb.close()
            csv_reader = iter(rows)  # Use iterator for consistency
        else:
            # Handle CSV file
            file_data = uploaded_file.read().decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(file_data))
        
        # Get column mappings from POST data
        column_mappings_json = request.POST.get('column_mappings', '{}')
        try:
            column_mappings = json.loads(column_mappings_json)
        except json.JSONDecodeError:
            column_mappings = {}
        
        # Get class assignment if provided
        assign_class_id = request.POST.get('assign_class', '') or None
        
        assign_class = None
        academic_year = None
        
        if assign_class_id:
            try:
                assign_class = Class.objects.get(pk=assign_class_id, school=school)
            except Class.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': 'Selected class not found.'
                }, status=400)
            
            # Automatically get current academic year for class assignment
            academic_year = SchoolInformation.get_current_academic_year(school=school)
            if not academic_year:
                return JsonResponse({
                    'success': False,
                    'error': 'No current academic year found. Please set a current academic year before importing students.'
                }, status=400)
        
        # Form and learning area removed - no longer part of import
        
        # Define field mappings (CSV column -> Student field)
        field_mappings = {
            'full_name': column_mappings.get('full_name', 'full_name'),
            'date_of_birth': column_mappings.get('date_of_birth', 'date_of_birth'),
            'gender': column_mappings.get('gender', 'gender'),
            'parent_contact': column_mappings.get('parent_contact', 'parent_contact'),
            'admission_date': column_mappings.get('admission_date', 'admission_date'),
            'email': column_mappings.get('email', 'email'),
            'admission_number': column_mappings.get('admission_number', 'admission_number'),
        }
        
        # Required fields - only full_name and gender (admission_number is optional, auto-generated if not provided)
        required_fields = ['full_name', 'gender']
        
        # Process rows
        imported = 0
        assigned_to_class = 0
        errors = []
        
        with transaction.atomic():
            for row_num, row in enumerate(csv_reader, start=2):  # Start at 2 (row 1 is header)
                try:
                    # Get data from row using column mappings
                    def get_field_value(field_name):
                        csv_column = field_mappings.get(field_name)
                        if csv_column and csv_column in row:
                            return row[csv_column].strip()
                        return ''
                    
                    full_name = get_field_value('full_name')
                    date_of_birth_str = get_field_value('date_of_birth')
                    date_of_birth = date_of_birth_str.strip() if date_of_birth_str else ''
                    gender_str = get_field_value('gender')
                    gender = gender_str.upper() if gender_str else ''
                    parent_contact_str = get_field_value('parent_contact')
                    parent_contact = parent_contact_str.strip() if parent_contact_str else ''
                    admission_date_str = get_field_value('admission_date')
                    admission_date = admission_date_str.strip() if admission_date_str else ''
                    email_str = get_field_value('email')
                    email = email_str.strip() if email_str else ''
                    admission_number_str = get_field_value('admission_number')
                    admission_number = admission_number_str.strip() if admission_number_str else ''
                    
                    # Validate required fields - only full_name and gender
                    if not full_name:
                        errors.append(f'Row {row_num}: Full name is required')
                        continue
                    
                    if not gender:
                        errors.append(f'Row {row_num}: Gender is required')
                        continue
                    
                    # Validate gender
                    if gender not in ['M', 'F']:
                        errors.append(f'Row {row_num}: Invalid gender (must be M or F)')
                        continue
                    
                    # Check if admission_number already exists (only if provided)
                    if admission_number:
                        if Student.objects.filter(admission_number=admission_number, school=school).exists():
                            errors.append(f'Row {row_num}: Admission number "{admission_number}" already exists')
                            continue
                    
                    # Provide defaults for required fields if not provided
                    if not date_of_birth:
                        # Default to 10 years ago from today (use January 1st to avoid leap year issues)
                        default_dob = date(date.today().year - 10, 1, 1)
                        date_of_birth = default_dob.strftime('%Y-%m-%d')
                    
                    if not parent_contact:
                        parent_contact = 'N/A'
                    
                    if not admission_date:
                        # Default to today's date
                        admission_date = date.today().strftime('%Y-%m-%d')
                    
                    # Parse date strings to date objects
                    try:
                        dob_date = datetime.strptime(date_of_birth, '%Y-%m-%d').date()
                    except (ValueError, TypeError):
                        errors.append(f'Row {row_num}: Invalid date of birth format (expected YYYY-MM-DD)')
                        continue
                    
                    try:
                        admission_date_obj = datetime.strptime(admission_date, '%Y-%m-%d').date()
                    except (ValueError, TypeError):
                        errors.append(f'Row {row_num}: Invalid admission date format (expected YYYY-MM-DD)')
                        continue
                    
                    # Create student - admission_number will be auto-generated by save() if not provided
                    student = Student(
                        full_name=full_name,
                        gender=gender,
                        date_of_birth=dob_date,
                        parent_contact=parent_contact,
                        admission_date=admission_date_obj,
                        school=school,
                    )
                    
                    # Set admission_number only if provided (otherwise it will be auto-generated)
                    if admission_number:
                        student.admission_number = admission_number
                    
                    # Set optional email if provided
                    if email:
                        student.email = email
                    else:
                        student.email = None
                    
                    # Save student with retry for auto-generated admission numbers (handle potential uniqueness conflicts)
                    max_retries = 5
                    saved = False
                    for attempt in range(max_retries):
                        try:
                            student.save()
                            saved = True
                            break
                        except IntegrityError as e:
                            # If admission_number conflict and it was auto-generated, try again with new ID
                            if 'admission_number' in str(e) and not admission_number:
                                # Reset admission_number to trigger regeneration
                                student.admission_number = None
                                continue
                            else:
                                # Other integrity error, re-raise
                                raise
                    
                    if not saved:
                        errors.append(f'Row {row_num}: Failed to generate unique admission number after {max_retries} attempts')
                        continue
                    
                    imported += 1
                    
                    # Assign to class if specified (academic_year is automatically set to current)
                    if assign_class:
                        # Check if student is already assigned to an active class
                        existing_assignment = StudentClass.objects.filter(
                            student=student,
                            is_active=True,
                            school=school
                        ).first()
                        
                        if existing_assignment:
                            # Deactivate existing assignment
                            existing_assignment.is_active = False
                            existing_assignment.save()
                        
                        # Create new assignment
                        student_class = StudentClass(
                            student=student,
                            assigned_class=assign_class,
                            assigned_by=request.user,
                            school=school,
                        )
                        student_class.save()
                        assigned_to_class += 1
                
                except Exception as e:
                    errors.append(f'Row {row_num}: {str(e)}')
        
        # Prepare response
        response_data = {
            'success': True,
            'imported': imported,
            'assigned_to_class': assigned_to_class,
            'errors': errors,
        }
        
        message_parts = []
        if imported > 0:
            message_parts.append(f'Successfully imported {imported} student(s)')
        else:
            message_parts.append('No students were imported.')
        
        if assigned_to_class > 0:
            message_parts.append(f'{assigned_to_class} student(s) assigned to class')
        
        response_data['message'] = '. '.join(message_parts) + '.'
        
        if errors:
            response_data['error_count'] = len(errors)
            response_data['error_summary'] = f'{len(errors)} error(s) occurred during import.'
        
        return JsonResponse(response_data)
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error processing import: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["GET"])
def student_profile_view(request):
    """
    View own student profile (for students to view their own profile).
    """
    if request.user.role != "student":
        messages.error(request, "You don't have permission to access this page.")
        return redirect("quiz_app:dashboard")
    
    if not request.user.student_profile:
        messages.error(request, "Student profile not found. Please contact administrator.")
        return redirect("quiz_app:dashboard")
    
    student = request.user.student_profile
    
    # Get student's enrollments
    enrollments = student.studentclassenrollment_set.filter(is_active=True).select_related(
        'assigned_class', 'academic_year', 'term'
    ).order_by('-academic_year__start_date', 'assigned_class__name')
    
    context = {
        'student': student,
        'enrollments': enrollments,
        'is_own_profile': True,
    }
    
    return render(request, 'student/student_detail.html', context)


@login_required
@require_http_methods(["GET"])
def student_detail_view(request, student_id):
    """
    View student details.
    
    For teachers: Only allows viewing students in classes they teach
    For admins/superadmins: Can view any student in their school
    """
    school = request.user.school
    
    # Handle teacher access
    if request.user.role == "teacher":
        if not hasattr(request.user, 'teacher_profile') or not request.user.teacher_profile:
            messages.error(request, "Teacher profile not found.")
            return redirect("quiz_app:dashboard")
        
        teacher = request.user.teacher_profile
        school = teacher.school
        
        # Get the student
        try:
            student = Student.objects.get(pk=student_id, school=school)
        except Student.DoesNotExist:
            messages.error(request, "Student not found.")
            return redirect("quiz_app:student_list")
        
        # Check if student is in a class the teacher teaches
        teacher_assignments = TeacherSubjectAssignment.objects.filter(
            teacher=teacher,
            is_active=True
        )
        if school:
            teacher_assignments = teacher_assignments.filter(school=school)
        
        teacher_classes = teacher_assignments.values_list('class_assigned', flat=True).distinct()
        
        # Check if student is assigned to any of the teacher's classes
        student_in_teacher_class = StudentClass.objects.filter(
            student=student,
            assigned_class__in=teacher_classes,
            is_active=True
        ).exists()
        
        if not student_in_teacher_class:
            messages.error(request, "You don't have permission to view this student. They are not in any of your classes.")
            return redirect("quiz_app:student_list")
    
    # Handle admin and superadmin access
    elif request.user.role == "superadmin":
        student = get_object_or_404(Student, pk=student_id)
    elif request.user.role == "admin":
        if not school:
            messages.error(request, "No school associated with your account.")
            return redirect("quiz_app:student_list")
        student = get_object_or_404(Student, pk=student_id, school=school)
    else:
        messages.error(request, "You don't have permission to access this page.")
        return redirect("quiz_app:dashboard")
    
    context = {
        'student': student,
    }
    
    return render(request, 'student/student_detail.html', context)

