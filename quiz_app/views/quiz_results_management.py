"""
Quiz Results Management views for teachers.

This module provides views for managing quiz results including:
- Listing quiz results/attempts
- Viewing individual quiz results
- Exporting results to Excel
- Printing individual result reports
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Q, Count, Avg, Max, Min
from django.utils import timezone
from io import BytesIO

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from ..models import (
    Quiz,
    QuizAttempt,
    QuizResponse,
    Teacher,
)


@login_required
@require_http_methods(["GET"])
def quiz_results_list_view(request):
    """
    Display list of quiz results/attempts for teacher's quizzes.
    """
    if request.user.role not in ["teacher", "admin"]:
        messages.error(request, "You don't have permission to access this page.")
        return redirect("quiz_app:dashboard")
    
    school = request.user.school
    
    # Teachers see only their quizzes, admins see all quizzes in school
    if request.user.role == "teacher":
        if not request.user.teacher_profile:
            messages.error(request, "Teacher profile not found. Please contact administrator.")
            return redirect("quiz_app:dashboard")
        teacher = request.user.teacher_profile
        teacher_quizzes = Quiz.objects.filter(teacher=teacher, school=school)
    else:
        # Admin sees all quizzes in their school
        teacher_quizzes = Quiz.objects.filter(school=school)
    
    # Get all submitted attempts
    all_attempts = QuizAttempt.objects.filter(
        quiz__in=teacher_quizzes,
        school=school,
        is_submitted=True
    ).select_related('quiz', 'student', 'academic_year', 'term')
    
    # Filter by quiz
    quiz_filter = request.GET.get('quiz', '')
    if quiz_filter:
        all_attempts = all_attempts.filter(quiz_id=quiz_filter)
    
    # Filter by academic year
    academic_year_filter = request.GET.get('academic_year', '')
    if academic_year_filter:
        all_attempts = all_attempts.filter(academic_year_id=academic_year_filter)
    
    # Filter by class (through student's current class)
    from ..models import StudentClass
    class_filter = request.GET.get('class', '')
    if class_filter:
        student_ids = StudentClass.objects.filter(
            assigned_class_id=class_filter,
            is_active=True,
            school=school
        ).values_list('student_id', flat=True)
        all_attempts = all_attempts.filter(student_id__in=student_ids)
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        all_attempts = all_attempts.filter(
            Q(student__full_name__icontains=search_query) |
            Q(student__admission_number__icontains=search_query) |
            Q(quiz__title__icontains=search_query)
        )
    
    # Get only best attempts per student per quiz
    # Use a more efficient approach: get max score per quiz-student pair, then get the attempt with that score
    from django.db.models import Max
    
    # Get the maximum score for each quiz-student combination
    max_scores = all_attempts.values('quiz', 'student').annotate(
        max_score=Max('score'),
        max_percentage=Max('percentage')
    )
    
    # Build a list of best attempt IDs
    best_attempt_ids = []
    for score_data in max_scores:
        # Get attempts with max score for this quiz-student pair
        student_quiz_attempts = all_attempts.filter(
            quiz_id=score_data['quiz'],
            student_id=score_data['student'],
            score=score_data['max_score']
        )
        
        # If multiple attempts have the same max score, pick the one with highest percentage, then most recent
        best_attempt = student_quiz_attempts.order_by('-percentage', '-submitted_at').first()
        if best_attempt:
            best_attempt_ids.append(best_attempt.id)
    
    # Filter to only best attempts
    attempts = QuizAttempt.objects.filter(
        id__in=best_attempt_ids
    ).select_related('quiz', 'student', 'academic_year', 'term').order_by('-submitted_at')
    
    # Get filter options (use all_attempts for filter options, not just best attempts)
    quizzes = teacher_quizzes.order_by('-created_at')
    from ..models import AcademicYear, Class
    academic_years = AcademicYear.objects.filter(
        id__in=all_attempts.values_list('academic_year_id', flat=True).distinct(),
        school=school
    ).order_by('-start_date')
    
    # Get classes from students who have attempts
    student_ids = all_attempts.values_list('student_id', flat=True).distinct()
    class_ids = StudentClass.objects.filter(
        student_id__in=student_ids,
        is_active=True,
        school=school
    ).values_list('assigned_class_id', flat=True).distinct()
    classes = Class.objects.filter(
        id__in=class_ids,
        school=school
    ).order_by('name')
    
    # Calculate statistics
    total_attempts = attempts.count()
    avg_score = attempts.aggregate(avg=Avg('score'))['avg'] or 0
    avg_percentage = attempts.aggregate(avg=Avg('percentage'))['avg'] or 0
    
    context = {
        'attempts': attempts,
        'quizzes': quizzes,
        'academic_years': academic_years,
        'classes': classes,
        'quiz_filter': quiz_filter,
        'academic_year_filter': academic_year_filter,
        'class_filter': class_filter,
        'search_query': search_query,
        'total_attempts': total_attempts,
        'avg_score': avg_score,
        'avg_percentage': avg_percentage,
    }
    
    return render(request, 'quiz/results/quiz_results_list.html', context)


@login_required
@require_http_methods(["GET"])
def quiz_result_detail_view(request, attempt_id):
    """
    Display detailed view of a quiz result/attempt.
    """
    if request.user.role not in ["teacher", "admin"]:
        messages.error(request, "You don't have permission to access this page.")
        return redirect("quiz_app:dashboard")
    
    if not request.user.teacher_profile:
        messages.error(request, "Teacher profile not found. Please contact administrator.")
        return redirect("quiz_app:dashboard")
    
    teacher = request.user.teacher_profile
    school = request.user.school
    
    attempt = get_object_or_404(
        QuizAttempt,
        pk=attempt_id,
        quiz__teacher=teacher,
        school=school,
        is_submitted=True
    )
    
    # Get all responses ordered by question order
    responses = attempt.responses.select_related(
        'question'
    ).prefetch_related(
        'question__answer_choices',
        'selected_choice',
        'selected_choices'
    ).order_by('question__order')
    
    # Check if print mode
    print_mode = request.GET.get('print', '') == '1'
    
    if print_mode:
        # Format data for unified template
        quiz = attempt.quiz
        all_questions = quiz.questions.all().order_by('order')
        answered_question_ids = set(responses.values_list('question_id', flat=True))
        school_info = school
        
        attempts = [{
            'attempt': attempt,
            'quiz': quiz,
            'responses': responses,
            'all_questions': all_questions,
            'answered_question_ids': answered_question_ids,
        }]
        
        context = {
            'attempts': attempts,
            'school_info': school_info,
            'is_bulk': False,
        }
        
        return render(request, 'quiz/results/unified_result_print.html', context)
    else:
        context = {
            'attempt': attempt,
            'responses': responses,
            'print_mode': print_mode,
        }
        return render(request, 'quiz/results/quiz_result_detail.html', context)


@login_required
@require_http_methods(["GET"])
def quiz_results_export_view(request):
    """
    Export quiz results to Excel.
    """
    if request.user.role not in ["teacher", "admin"]:
        messages.error(request, "You don't have permission to perform this action.")
        return redirect("quiz_app:dashboard")
    
    if not request.user.teacher_profile:
        messages.error(request, "Teacher profile not found. Please contact administrator.")
        return redirect("quiz_app:dashboard")
    
    if not OPENPYXL_AVAILABLE:
        messages.error(request, "Excel export requires openpyxl. Please install it.")
        return redirect("quiz_app:quiz_results_list")
    
    teacher = request.user.teacher_profile
    school = request.user.school
    
    # Get quizzes created by this teacher
    teacher_quizzes = Quiz.objects.filter(teacher=teacher, school=school)
    
    # Get filter parameters
    quiz_id = request.GET.get('quiz', '')
    academic_year_id = request.GET.get('academic_year', '')
    class_id = request.GET.get('class', '')
    
    # Build query - get all attempts first
    all_attempts = QuizAttempt.objects.filter(
        quiz__in=teacher_quizzes,
        school=school,
        is_submitted=True
    ).select_related('quiz', 'student', 'academic_year', 'term')
    
    if quiz_id:
        all_attempts = all_attempts.filter(quiz_id=quiz_id)
    if academic_year_id:
        all_attempts = all_attempts.filter(academic_year_id=academic_year_id)
    if class_id:
        from ..models import StudentClass
        student_ids = StudentClass.objects.filter(
            assigned_class_id=class_id,
            is_active=True,
            school=school
        ).values_list('student_id', flat=True)
        all_attempts = all_attempts.filter(student_id__in=student_ids)
    
    # Get only best attempts per student per quiz
    from django.db.models import Max
    
    # Get the maximum score for each quiz-student combination
    max_scores = all_attempts.values('quiz', 'student').annotate(
        max_score=Max('score'),
        max_percentage=Max('percentage')
    )
    
    # Build a list of best attempt IDs
    best_attempt_ids = []
    for score_data in max_scores:
        # Get attempts with max score for this quiz-student pair
        student_quiz_attempts = all_attempts.filter(
            quiz_id=score_data['quiz'],
            student_id=score_data['student'],
            score=score_data['max_score']
        )
        
        # If multiple attempts have the same max score, pick the one with highest percentage, then most recent
        best_attempt = student_quiz_attempts.order_by('-percentage', '-submitted_at').first()
        if best_attempt:
            best_attempt_ids.append(best_attempt.id)
    
    # Filter to only best attempts
    attempts = QuizAttempt.objects.filter(
        id__in=best_attempt_ids
    ).select_related('quiz', 'student', 'academic_year', 'term').order_by('quiz__title', 'student__full_name', '-submitted_at')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Quiz Results"
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'Student Name',
        'Admission Number',
        'Quiz Title',
        'Subject',
        'Attempt #',
        'Score',
        'Total Marks',
        'Percentage',
        'Correct Answers',
        'Wrong Answers',
        'Unanswered',
        'Started At',
        'Submitted At',
        'Time Taken',
        'Academic Year',
        'Term',
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data
    for row_num, attempt in enumerate(attempts, 2):
        time_taken_str = ""
        if attempt.time_taken:
            total_seconds = int(attempt.time_taken.total_seconds())
            hours = total_seconds // 3600
            minutes = (total_seconds % 3600) // 60
            seconds = total_seconds % 60
            if hours > 0:
                time_taken_str = f"{hours}h {minutes}m {seconds}s"
            elif minutes > 0:
                time_taken_str = f"{minutes}m {seconds}s"
            else:
                time_taken_str = f"{seconds}s"
        
        row_data = [
            attempt.student.full_name,
            attempt.student.admission_number,
            attempt.quiz.title,
            attempt.quiz.subject.subject_name if attempt.quiz.subject else '',
            attempt.attempt_number,
            attempt.score,
            attempt.quiz.total_marks,
            attempt.percentage,
            attempt.correct_answers,
            attempt.wrong_answers,
            attempt.unanswered,
            attempt.started_at.strftime('%Y-%m-%d %H:%M:%S') if attempt.started_at else '',
            attempt.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if attempt.submitted_at else '',
            time_taken_str,
            attempt.academic_year.name if attempt.academic_year else '',
            attempt.term.name if attempt.term else '',
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            if col_num in [6, 7, 8]:  # Score, Total Marks, Percentage columns
                cell.alignment = Alignment(horizontal="center")
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"quiz_results_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response.write(output.getvalue())
    output.close()
    
    return response


@login_required
@require_http_methods(["GET"])
def quiz_result_print_view(request, attempt_id):
    """
    Print-friendly view of a quiz result.
    """
    if request.user.role not in ["teacher", "admin"]:
        messages.error(request, "You don't have permission to access this page.")
        return redirect("quiz_app:dashboard")
    
    if not request.user.teacher_profile:
        messages.error(request, "Teacher profile not found. Please contact administrator.")
        return redirect("quiz_app:dashboard")
    
    teacher = request.user.teacher_profile
    school = request.user.school
    
    attempt = get_object_or_404(
        QuizAttempt,
        pk=attempt_id,
        quiz__teacher=teacher,
        school=school,
        is_submitted=True
    )
    
    # Get all responses ordered by question order
    responses = attempt.responses.select_related(
        'question'
    ).prefetch_related(
        'question__answer_choices',
        'selected_choice',
        'selected_choices'
    ).order_by('question__order')
    
    # Format data for unified template
    quiz = attempt.quiz
    all_questions = quiz.questions.all().order_by('order')
    answered_question_ids = set(responses.values_list('question_id', flat=True))
    school_info = school
    
    attempts = [{
        'attempt': attempt,
        'quiz': quiz,
        'responses': responses,
        'all_questions': all_questions,
        'answered_question_ids': answered_question_ids,
    }]
    
    context = {
        'attempts': attempts,
        'school_info': school_info,
        'is_bulk': False,
    }
    
    return render(request, 'quiz/results/unified_result_print.html', context)


@login_required
@require_http_methods(["GET"])
def quiz_results_bulk_print_view(request):
    """
    Bulk print view for multiple quiz results.
    Allows teachers to print all student results at once.
    """
    if request.user.role not in ["teacher", "admin"]:
        messages.error(request, "You don't have permission to access this page.")
        return redirect("quiz_app:dashboard")
    
    school = request.user.school
    
    # Teachers see only their quizzes, admins see all quizzes in school
    if request.user.role == "teacher":
        if not request.user.teacher_profile:
            messages.error(request, "Teacher profile not found. Please contact administrator.")
            return redirect("quiz_app:dashboard")
        teacher = request.user.teacher_profile
        teacher_quizzes = Quiz.objects.filter(teacher=teacher, school=school)
    else:
        # Admin sees all quizzes in their school
        teacher_quizzes = Quiz.objects.filter(school=school)
    
    # Get attempt IDs from query parameters
    attempt_ids = request.GET.getlist('attempt_ids')
    
    if not attempt_ids:
        # If no specific attempts selected, use filters from query params
        all_attempts = QuizAttempt.objects.filter(
            quiz__in=teacher_quizzes,
            school=school,
            is_submitted=True
        ).select_related('quiz', 'student', 'academic_year', 'term')
        
        # Apply filters
        quiz_filter = request.GET.get('quiz', '')
        if quiz_filter:
            all_attempts = all_attempts.filter(quiz_id=quiz_filter)
        
        academic_year_filter = request.GET.get('academic_year', '')
        if academic_year_filter:
            all_attempts = all_attempts.filter(academic_year_id=academic_year_filter)
        
        from ..models import StudentClass
        class_filter = request.GET.get('class', '')
        if class_filter:
            student_ids = StudentClass.objects.filter(
                assigned_class_id=class_filter,
                is_active=True,
                school=school
            ).values_list('student_id', flat=True)
            all_attempts = all_attempts.filter(student_id__in=student_ids)
        
        search_query = request.GET.get('search', '')
        if search_query:
            all_attempts = all_attempts.filter(
                Q(student__full_name__icontains=search_query) |
                Q(student__admission_number__icontains=search_query) |
                Q(quiz__title__icontains=search_query)
            )
        
        # Get only best attempts per student per quiz
        from django.db.models import Max
        
        max_scores = all_attempts.values('quiz', 'student').annotate(
            max_score=Max('score'),
            max_percentage=Max('percentage')
        )
        
        best_attempt_ids = []
        for score_data in max_scores:
            student_quiz_attempts = all_attempts.filter(
                quiz_id=score_data['quiz'],
                student_id=score_data['student'],
                score=score_data['max_score']
            )
            best_attempt = student_quiz_attempts.order_by('-percentage', '-submitted_at').first()
            if best_attempt:
                best_attempt_ids.append(best_attempt.id)
        
        attempts_queryset = QuizAttempt.objects.filter(
            id__in=best_attempt_ids
        ).select_related('quiz', 'student', 'academic_year', 'term').order_by('quiz__title', 'student__full_name')
    else:
        # Get specific attempts
        attempts_queryset = QuizAttempt.objects.filter(
            id__in=attempt_ids,
            quiz__in=teacher_quizzes,
            school=school,
            is_submitted=True
        ).select_related('quiz', 'student', 'academic_year', 'term').order_by('quiz__title', 'student__full_name')
    
    # Prepare data for unified template
    attempts_data = []
    for attempt in attempts_queryset:
        quiz = attempt.quiz
        
        # Get all responses ordered by question order
        responses = attempt.responses.select_related(
            'question'
        ).prefetch_related(
            'question__answer_choices',
            'selected_choice',
            'selected_choices'
        ).order_by('question__order')
        
        # Get all questions to show unanswered ones
        all_questions = quiz.questions.all().order_by('order')
        answered_question_ids = set(responses.values_list('question_id', flat=True))
        
        attempts_data.append({
            'attempt': attempt,
            'quiz': quiz,
            'responses': responses,
            'all_questions': all_questions,
            'answered_question_ids': answered_question_ids,
        })
    
    if not attempts_data:
        messages.warning(request, "No results found to print.")
        return redirect("quiz_app:quiz_results_list")
    
    school_info = school
    
    context = {
        'attempts': attempts_data,
        'school_info': school_info,
        'is_bulk': True,
    }
    
    return render(request, 'quiz/results/unified_result_print.html', context)

