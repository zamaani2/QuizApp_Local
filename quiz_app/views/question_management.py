"""
Question management views for teachers.

This module provides views for managing questions within quizzes including:
- Listing questions for a quiz
- Creating new questions
- Editing existing questions
- Deleting questions
- Reordering questions
- Bulk importing questions
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Q, Max
from django.db import transaction
from django.core.exceptions import ValidationError
import json
import csv
import io

from ..models import (
    Quiz,
    Question,
    AnswerChoice,
    Teacher,
    SchoolInformation,
)


@login_required
@require_http_methods(["GET"])
def question_list_view(request, quiz_id):
    """
    List all questions for a quiz.
    This is typically shown in the quiz detail page, but can be a separate view.
    """
    if request.user.role not in ["teacher", "admin"]:
        messages.error(request, "You don't have permission to access this page.")
        return redirect("quiz_app:dashboard")
    
    school = request.user.school
    
    # Teachers can only access their own quizzes, admins can access any quiz in their school
    if request.user.role == "teacher":
        if not request.user.teacher_profile:
            messages.error(request, "Teacher profile not found. Please contact administrator.")
            return redirect("quiz_app:dashboard")
        teacher = request.user.teacher_profile
        quiz = get_object_or_404(Quiz, pk=quiz_id, teacher=teacher, school=school)
    else:
        # Admin can access any quiz in their school
        quiz = get_object_or_404(Quiz, pk=quiz_id, school=school)
        teacher = quiz.teacher  # Use quiz's teacher for context
    
    # Get questions with answer choices
    questions = quiz.questions.all().prefetch_related('answer_choices').order_by('order', 'created_at')
    
    context = {
        'quiz': quiz,
        'questions': questions,
    }
    
    return render(request, 'quiz/question/question_list.html', context)


@login_required
@require_http_methods(["GET", "POST"])
def question_create_view(request, quiz_id):
    """
    Create a new question for a quiz.
    
    GET: Returns question creation form
    POST: Creates question and returns JSON response or redirects
    """
    if request.user.role not in ["teacher", "admin"]:
        if request.method == "GET":
            return JsonResponse({'error': 'Permission denied'}, status=403)
        messages.error(request, "You don't have permission to perform this action.")
        return redirect("quiz_app:quiz_list")
    
    school = request.user.school
    
    # Teachers can only access their own quizzes, admins can access any quiz in their school
    if request.user.role == "teacher":
        if not request.user.teacher_profile:
            if request.method == "GET":
                return JsonResponse({'error': 'Teacher profile not found'}, status=404)
            messages.error(request, "Teacher profile not found. Please contact administrator.")
            return redirect("quiz_app:quiz_list")
        teacher = request.user.teacher_profile
        quiz = get_object_or_404(Quiz, pk=quiz_id, teacher=teacher, school=school)
    else:
        # Admin can access any quiz in their school
        quiz = get_object_or_404(Quiz, pk=quiz_id, school=school)
        teacher = quiz.teacher  # Use quiz's teacher for context
    
    if request.method == "GET":
        # Check if this is an AJAX request for modal
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            html = render(request, 'quiz/question/partials/question_form.html', {
                'quiz': quiz,
                'question': None,
            }).content.decode('utf-8')
            return JsonResponse({'html': html})
        
        # Regular page request
        return render(request, 'quiz/question/question_create.html', {
            'quiz': quiz,
        })
    
    # POST - Create question
    try:
        with transaction.atomic():
            # Get form data
            question_type = request.POST.get('question_type', 'multiple_choice')
            question_text = request.POST.get('question_text', '').strip()
            marks = request.POST.get('marks', '1')
            order = request.POST.get('order', '')
            explanation = request.POST.get('explanation', '').strip() or None
            difficulty = request.POST.get('difficulty', 'medium')
            negative_marks = request.POST.get('negative_marks', '0')
            is_required = request.POST.get('is_required', 'on') == 'on'
            
            # Validation
            if not question_text:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'error': 'Question text is required.'
                    }, status=400)
                messages.error(request, 'Question text is required.')
                return redirect("quiz_app:question_create", quiz_id=quiz_id)
            
            # Get next order if not provided
            if not order:
                max_order = Question.objects.filter(quiz=quiz).aggregate(
                    max_order=Max('order')
                )['max_order'] or 0
                order = max_order + 1
            else:
                order = int(order)
            
            # Create question
            question = Question(
                quiz=quiz,
                question_type=question_type,
                question_text=question_text,
                marks=int(marks) if marks else 1,
                order=order,
                explanation=explanation,
                difficulty=difficulty,
                negative_marks=float(negative_marks) if negative_marks else 0.0,
                is_required=is_required,
                school=school,
            )
            question.save()
            
            # Handle question type specific data
            if question_type in ['multiple_choice', 'true_false']:
                # Get answer choices
                choice_texts = request.POST.getlist('choice_text')
                is_correct_list = request.POST.getlist('is_correct')
                
                # For true/false, create default choices if not provided
                if question_type == 'true_false' and not choice_texts:
                    choice_texts = ['True', 'False']
                    # Get correct answer
                    tf_correct = request.POST.get('tf_correct', 'true')
                    is_correct_list = ['0'] if tf_correct == 'true' else ['1']
                
                # Create answer choices
                for idx, choice_text in enumerate(choice_texts):
                    if choice_text.strip():
                        is_correct = str(idx) in is_correct_list
                        AnswerChoice.objects.create(
                            question=question,
                            choice_text=choice_text.strip(),
                            is_correct=is_correct,
                            order=idx + 1,
                            school=school,
                        )
            
            elif question_type in ['short_answer', 'fill_blank']:
                # Get correct answer text
                correct_answer = request.POST.get('correct_answer', '').strip()
                if correct_answer:
                    question.correct_answer_text = correct_answer
                    question.save()
            
            # Recalculate quiz total marks
            quiz.calculate_total_marks()
            
            # Check if "Save & Add Another" was clicked
            save_and_add = request.POST.get('save_and_add') or request.POST.get('add_another')
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': f'Question created successfully.',
                    'question_id': question.id,
                    'redirect': None if save_and_add else f'/quizzes/{quiz_id}/detail/',
                })
            
            if save_and_add:
                messages.success(request, 'Question created successfully. Add another question below.')
                return redirect("quiz_app:question_create", quiz_id=quiz_id)
            else:
                messages.success(request, 'Question created successfully.')
                return redirect("quiz_app:quiz_detail", quiz_id=quiz_id)
    
    except Exception as e:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': f'Error creating question: {str(e)}'
            }, status=500)
        messages.error(request, f'Error creating question: {str(e)}')
        return redirect("quiz_app:question_create", quiz_id=quiz_id)


@login_required
@require_http_methods(["GET", "POST"])
def question_edit_view(request, quiz_id, question_id):
    """
    Edit an existing question.
    
    GET: Returns question edit form
    POST: Updates question and returns JSON response or redirects
    """
    if request.user.role not in ["teacher", "admin"]:
        if request.method == "GET":
            return JsonResponse({'error': 'Permission denied'}, status=403)
        messages.error(request, "You don't have permission to perform this action.")
        return redirect("quiz_app:quiz_list")
    
    school = request.user.school
    
    # Teachers can only access their own quizzes, admins can access any quiz in their school
    if request.user.role == "teacher":
        if not request.user.teacher_profile:
            if request.method == "GET":
                return JsonResponse({'error': 'Teacher profile not found'}, status=404)
            messages.error(request, "Teacher profile not found. Please contact administrator.")
            return redirect("quiz_app:quiz_list")
        teacher = request.user.teacher_profile
        quiz = get_object_or_404(Quiz, pk=quiz_id, teacher=teacher, school=school)
    else:
        # Admin can access any quiz in their school
        quiz = get_object_or_404(Quiz, pk=quiz_id, school=school)
        teacher = quiz.teacher  # Use quiz's teacher for context
    question = get_object_or_404(Question, pk=question_id, quiz=quiz, school=school)
    
    if request.method == "GET":
        # Get answer choices for multiple choice/true false
        answer_choices = []
        tf_correct_answer = 'true'  # Default for True/False
        
        if question.question_type in ['multiple_choice', 'true_false']:
            answer_choices = list(question.answer_choices.all().order_by('order'))
            
            # For True/False questions, ensure we have both True and False choices
            if question.question_type == 'true_false':
                true_choice = None
                false_choice = None
                
                # Find existing True/False choices
                for choice in answer_choices:
                    if choice.choice_text == 'True':
                        true_choice = choice
                        if choice.is_correct:
                            tf_correct_answer = 'true'
                    elif choice.choice_text == 'False':
                        false_choice = choice
                        if choice.is_correct:
                            tf_correct_answer = 'false'
                
                # Create True choice if it doesn't exist
                if not true_choice:
                    true_choice = AnswerChoice.objects.create(
                        question=question,
                        choice_text='True',
                        is_correct=False,  # Will be set based on existing correct answer
                        order=1,
                        school=school,
                    )
                    answer_choices.insert(0, true_choice)
                
                # Create False choice if it doesn't exist
                if not false_choice:
                    false_choice = AnswerChoice.objects.create(
                        question=question,
                        choice_text='False',
                        is_correct=False,  # Will be set based on existing correct answer
                        order=2,
                        school=school,
                    )
                    answer_choices.append(false_choice)
                
                # Determine correct answer from existing choices
                # If neither is correct, default to True
                found_correct = False
                for choice in answer_choices:
                    if choice.is_correct:
                        if choice.choice_text == 'True':
                            tf_correct_answer = 'true'
                        else:
                            tf_correct_answer = 'false'
                        found_correct = True
                        break
                
                # If no correct answer found, ensure at least one is marked correct (default to True)
                if not found_correct and true_choice:
                    true_choice.is_correct = True
                    true_choice.save()
                    if false_choice:
                        false_choice.is_correct = False
                        false_choice.save()
                    tf_correct_answer = 'true'
                
                # Re-order choices to ensure True is first, False is second
                answer_choices.sort(key=lambda x: (x.choice_text != 'True', x.order))
        
        # Check if this is an AJAX request for modal
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            html = render(request, 'quiz/question/partials/question_form.html', {
                'quiz': quiz,
                'question': question,
                'answer_choices': answer_choices,
                'tf_correct_answer': tf_correct_answer if question.question_type == 'true_false' else None,
            }).content.decode('utf-8')
            return JsonResponse({'html': html})
        
        # Regular page request
        return render(request, 'quiz/question/question_edit.html', {
            'quiz': quiz,
            'question': question,
            'answer_choices': answer_choices,
            'tf_correct_answer': tf_correct_answer if question.question_type == 'true_false' else None,
        })
    
    # POST - Update question
    try:
        with transaction.atomic():
            # Get form data
            question.question_text = request.POST.get('question_text', '').strip()
            question.marks = int(request.POST.get('marks', '1')) if request.POST.get('marks') else 1
            question.order = int(request.POST.get('order', question.order)) if request.POST.get('order') else question.order
            question.explanation = request.POST.get('explanation', '').strip() or None
            question.difficulty = request.POST.get('difficulty', 'medium')
            question.negative_marks = float(request.POST.get('negative_marks', '0')) if request.POST.get('negative_marks') else 0.0
            question.is_required = request.POST.get('is_required', 'on') == 'on'
            
            # Validation
            if not question.question_text:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'error': 'Question text is required.'
                    }, status=400)
                messages.error(request, 'Question text is required.')
                return redirect("quiz_app:question_edit", quiz_id=quiz_id, question_id=question_id)
            
            question.save()
            
            # Handle question type specific data
            if question.question_type == 'true_false':
                # Handle True/False questions - they use radio buttons with name="tf_correct"
                # Get the selected True/False value from POST data
                tf_correct = request.POST.get('tf_correct', '').strip().lower()
                
                # Debug logging (remove in production)
                import logging
                logger = logging.getLogger(__name__)
                logger.debug(f"True/False question update - tf_correct from POST: {tf_correct}")
                logger.debug(f"POST data keys: {list(request.POST.keys())}")
                logger.debug(f"All POST data: {dict(request.POST)}")
                
                # Validate tf_correct value
                if tf_correct not in ['true', 'false']:
                    # If not provided or invalid, check existing choices
                    existing_correct = question.answer_choices.filter(is_correct=True).first()
                    if existing_correct:
                        tf_correct = 'true' if existing_correct.choice_text.strip() == 'True' else 'false'
                    else:
                        tf_correct = 'true'  # Default to True
                    logger.debug(f"tf_correct determined from existing: {tf_correct}")
                
                # Get or create True/False choices using filter for efficiency
                true_choice = question.answer_choices.filter(choice_text='True').first()
                false_choice = question.answer_choices.filter(choice_text='False').first()
                
                # Create or update True choice
                if not true_choice:
                    true_choice = AnswerChoice.objects.create(
                        question=question,
                        choice_text='True',
                        is_correct=(tf_correct == 'true'),
                        order=1,
                        school=school,
                    )
                    logger.debug(f"Created True choice with is_correct={true_choice.is_correct}")
                else:
                    true_choice.is_correct = (tf_correct == 'true')
                    true_choice.order = 1
                    true_choice.save()
                    logger.debug(f"Updated True choice with is_correct={true_choice.is_correct}")
                
                # Create or update False choice
                if not false_choice:
                    false_choice = AnswerChoice.objects.create(
                        question=question,
                        choice_text='False',
                        is_correct=(tf_correct == 'false'),
                        order=2,
                        school=school,
                    )
                    logger.debug(f"Created False choice with is_correct={false_choice.is_correct}")
                else:
                    false_choice.is_correct = (tf_correct == 'false')
                    false_choice.order = 2
                    false_choice.save()
                    logger.debug(f"Updated False choice with is_correct={false_choice.is_correct}")
                
                # Delete any other choices that shouldn't exist
                question.answer_choices.exclude(
                    pk__in=[true_choice.id, false_choice.id]
                ).delete()
            
            elif question.question_type == 'multiple_choice':
                # Handle Multiple Choice questions
                # Get existing choices
                existing_choices = {choice.id: choice for choice in question.answer_choices.all()}
                
                # Get form data
                choice_ids = request.POST.getlist('choice_id')
                choice_texts = request.POST.getlist('choice_text')
                is_correct_list = request.POST.getlist('is_correct')
                
                # Debug logging
                import logging
                logger = logging.getLogger(__name__)
                logger.debug(f"Multiple choice update - choice_ids: {choice_ids}, choice_texts: {choice_texts}, is_correct_list: {is_correct_list}")
                
                # Ensure we have matching lengths - zip stops at shortest, so we need to handle this carefully
                max_len = max(len(choice_ids), len(choice_texts))
                
                # Pad shorter lists with empty strings
                while len(choice_ids) < max_len:
                    choice_ids.append('')
                while len(choice_texts) < max_len:
                    choice_texts.append('')
                
                # Update or create choices
                processed_choice_ids = set()
                choice_index = 0  # Track actual index for choices with text
                
                for choice_id, choice_text in zip(choice_ids, choice_texts):
                    if not choice_text or not choice_text.strip():
                        # Skip empty choice texts
                        continue
                    
                    choice_text_clean = choice_text.strip()
                    # Check if this choice index is marked as correct
                    is_correct = str(choice_index) in is_correct_list
                    
                    # Try to find existing choice by ID
                    choice_id_clean = choice_id.strip() if choice_id else ''
                    
                    if choice_id_clean and choice_id_clean.isdigit():
                        choice_id_int = int(choice_id_clean)
                        if choice_id_int in existing_choices:
                            # Update existing choice
                            choice = existing_choices[choice_id_int]
                            choice.choice_text = choice_text_clean
                            choice.is_correct = is_correct
                            choice.order = choice_index + 1
                            choice.save()
                            processed_choice_ids.add(choice_id_int)
                            logger.debug(f"Updated choice {choice_id_int}: {choice_text_clean}, is_correct={is_correct}")
                            choice_index += 1
                            continue
                    
                    # Create new choice if ID not found or invalid
                    new_choice = AnswerChoice.objects.create(
                        question=question,
                        choice_text=choice_text_clean,
                        is_correct=is_correct,
                        order=choice_index + 1,
                        school=school,
                    )
                    processed_choice_ids.add(new_choice.id)
                    logger.debug(f"Created new choice: {choice_text_clean}, is_correct={is_correct}")
                    choice_index += 1
                
                # Delete choices that were not processed (removed from form)
                deleted_count = 0
                for choice_id, choice in existing_choices.items():
                    if choice_id not in processed_choice_ids:
                        choice.delete()
                        deleted_count += 1
                
                if deleted_count > 0:
                    logger.debug(f"Deleted {deleted_count} choice(s) that were removed from form")
            
            elif question.question_type in ['short_answer', 'fill_blank']:
                # Update correct answer text
                correct_answer = request.POST.get('correct_answer', '').strip()
                question.correct_answer_text = correct_answer if correct_answer else None
                question.save()
            
            # Recalculate quiz total marks
            quiz.calculate_total_marks()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': f'Question updated successfully.',
                    'question_id': question.id,
                })
            
            messages.success(request, 'Question updated successfully.')
            return redirect("quiz_app:quiz_detail", quiz_id=quiz_id)
    
    except Exception as e:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': f'Error updating question: {str(e)}'
            }, status=500)
        messages.error(request, f'Error updating question: {str(e)}')
        return redirect("quiz_app:question_edit", quiz_id=quiz_id, question_id=question_id)


@login_required
@require_http_methods(["POST"])
def question_delete_view(request, quiz_id, question_id):
    """
    Delete a question.
    """
    if request.user.role not in ["teacher", "admin"]:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if not request.user.teacher_profile:
        return JsonResponse({'error': 'Teacher profile not found'}, status=404)
    
    teacher = request.user.teacher_profile
    school = request.user.school
    quiz = get_object_or_404(Quiz, pk=quiz_id, teacher=teacher, school=school)
    question = get_object_or_404(Question, pk=question_id, quiz=quiz, school=school)
    
    try:
        question_text = question.question_text[:50]
        question.delete()
        
        # Recalculate quiz total marks
        quiz.calculate_total_marks()
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': f'Question deleted successfully.'
            })
        
        messages.success(request, 'Question deleted successfully.')
        return redirect("quiz_app:quiz_detail", quiz_id=quiz_id)
    
    except Exception as e:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': f'Error deleting question: {str(e)}'
            }, status=500)
        messages.error(request, f'Error deleting question: {str(e)}')
        return redirect("quiz_app:quiz_detail", quiz_id=quiz_id)


@login_required
@require_http_methods(["POST"])
def question_reorder_view(request, quiz_id):
    """
    Reorder questions in a quiz (AJAX endpoint).
    """
    if request.user.role not in ["teacher", "admin"]:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if not request.user.teacher_profile:
        return JsonResponse({'error': 'Teacher profile not found'}, status=404)
    
    teacher = request.user.teacher_profile
    school = request.user.school
    quiz = get_object_or_404(Quiz, pk=quiz_id, teacher=teacher, school=school)
    
    try:
        data = json.loads(request.body)
        question_orders = data.get('question_orders', [])  # List of {question_id: order}
        
        with transaction.atomic():
            for item in question_orders:
                question_id = item.get('question_id')
                order = item.get('order')
                if question_id and order:
                    Question.objects.filter(
                        pk=question_id,
                        quiz=quiz,
                        school=school
                    ).update(order=order)
        
        return JsonResponse({
            'success': True,
            'message': 'Questions reordered successfully.'
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error reordering questions: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["GET", "POST"])
def question_duplicate_view(request, quiz_id, question_id):
    """
    Duplicate a question.
    """
    if request.user.role not in ["teacher", "admin"]:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if not request.user.teacher_profile:
        return JsonResponse({'error': 'Teacher profile not found'}, status=404)
    
    teacher = request.user.teacher_profile
    school = request.user.school
    quiz = get_object_or_404(Quiz, pk=quiz_id, teacher=teacher, school=school)
    question = get_object_or_404(Question, pk=question_id, quiz=quiz, school=school)
    
    try:
        with transaction.atomic():
            # Get next order
            max_order = Question.objects.filter(quiz=quiz).aggregate(
                max_order=Max('order')
            )['max_order'] or 0
            
            # Duplicate question
            new_question = Question(
                quiz=quiz,
                question_type=question.question_type,
                question_text=question.question_text,
                marks=question.marks,
                order=max_order + 1,
                explanation=question.explanation,
                difficulty=question.difficulty,
                negative_marks=question.negative_marks,
                is_required=question.is_required,
                correct_answer_text=question.correct_answer_text,
                school=school,
            )
            new_question.save()
            
            # Duplicate answer choices if any
            if question.question_type in ['multiple_choice', 'true_false']:
                for choice in question.answer_choices.all():
                    AnswerChoice.objects.create(
                        question=new_question,
                        choice_text=choice.choice_text,
                        is_correct=choice.is_correct,
                        order=choice.order,
                        explanation=choice.explanation,
                        partial_credit=choice.partial_credit,
                        school=school,
                    )
            
            # Recalculate quiz total marks
            quiz.calculate_total_marks()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Question duplicated successfully.',
                    'question_id': new_question.id,
                })
            
            messages.success(request, 'Question duplicated successfully.')
            return redirect("quiz_app:quiz_detail", quiz_id=quiz_id)
    
    except Exception as e:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': f'Error duplicating question: {str(e)}'
            }, status=500)
        messages.error(request, f'Error duplicating question: {str(e)}')
        return redirect("quiz_app:quiz_detail", quiz_id=quiz_id)


@login_required
@require_http_methods(["GET", "POST"])
def question_bulk_import_view(request, quiz_id):
    """
    Bulk import questions from CSV/Excel file.
    
    GET: Returns import form
    POST: Processes import file
    """
    if request.user.role not in ["teacher", "admin"]:
        if request.method == "GET":
            return JsonResponse({'error': 'Permission denied'}, status=403)
        messages.error(request, "You don't have permission to perform this action.")
        return redirect("quiz_app:quiz_list")
    
    school = request.user.school
    
    # Teachers can only access their own quizzes, admins can access any quiz in their school
    if request.user.role == "teacher":
        if not request.user.teacher_profile:
            if request.method == "GET":
                return JsonResponse({'error': 'Teacher profile not found'}, status=404)
            messages.error(request, "Teacher profile not found. Please contact administrator.")
            return redirect("quiz_app:quiz_list")
        teacher = request.user.teacher_profile
        quiz = get_object_or_404(Quiz, pk=quiz_id, teacher=teacher, school=school)
    else:
        # Admin can access any quiz in their school
        quiz = get_object_or_404(Quiz, pk=quiz_id, school=school)
        teacher = quiz.teacher  # Use quiz's teacher for context
    
    if request.method == "GET":
        # Check if this is an AJAX request for modal
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            html = render(request, 'quiz/question/partials/bulk_import_modal.html', {
                'quiz': quiz,
            }).content.decode('utf-8')
            return JsonResponse({'html': html})
        
        # Regular page request
        return render(request, 'quiz/question/question_import.html', {
            'quiz': quiz,
        })
    
    # POST - Process import
    try:
        if 'import_file' not in request.FILES:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'error': 'No file uploaded.'
                }, status=400)
            messages.error(request, 'No file uploaded.')
            return redirect("quiz_app:question_bulk_import", quiz_id=quiz_id)
        
        import_file = request.FILES['import_file']
        file_name = import_file.name.lower()
        
        # Determine file format from extension
        if file_name.endswith(('.xlsx', '.xls')):
            # Excel file
            try:
                imported, errors = _import_questions_from_excel(quiz, import_file, school)
            except ImportError:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'error': 'Excel support requires openpyxl. Please install it or use CSV format.'
                    }, status=400)
                messages.error(request, 'Excel support requires openpyxl. Please install it or use CSV format.')
                return redirect("quiz_app:question_bulk_import", quiz_id=quiz_id)
        elif file_name.endswith('.csv'):
            # CSV file
            imported, errors = _import_questions_from_csv(quiz, import_file, school)
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'error': 'Unsupported file format. Please use CSV or Excel (.xlsx, .xls) files.'
                }, status=400)
            messages.error(request, 'Unsupported file format. Please use CSV or Excel (.xlsx, .xls) files.')
            return redirect("quiz_app:question_bulk_import", quiz_id=quiz_id)
        
        # Recalculate quiz total marks
        quiz.calculate_total_marks()
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            response_data = {
                'success': True,
                'imported': imported,
                'errors': errors,
            }
            if imported > 0:
                response_data['message'] = f'Successfully imported {imported} question(s).'
            else:
                response_data['message'] = 'No questions were imported.'
            if errors:
                response_data['error_count'] = len(errors)
            return JsonResponse(response_data)
        
        if imported > 0:
            messages.success(request, f'Successfully imported {imported} question(s).')
        if errors:
            messages.warning(request, f'{len(errors)} error(s) occurred during import.')
        
        return redirect("quiz_app:quiz_detail", quiz_id=quiz_id)
    
    except Exception as e:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': f'Error processing import: {str(e)}'
            }, status=500)
        messages.error(request, f'Error processing import: {str(e)}')
        return redirect("quiz_app:question_bulk_import", quiz_id=quiz_id)


def _import_questions_from_csv(quiz, csv_file, school):
    """
    Helper function to import questions from CSV file.
    
    Expected CSV format:
    - question_text (required)
    - question_type (optional, default: multiple_choice)
    - marks (optional, default: 1)
    - difficulty (optional, default: medium)
    - correct_answer (required for multiple_choice/true_false: index or 'true'/'false')
    - choice1, choice2, choice3, choice4 (optional, for multiple_choice)
    - explanation (optional)
    
    Returns: (imported_count, errors_list)
    """
    
    imported = 0
    errors = []
    
    try:
        # Try different encodings
        try:
            file_data = csv_file.read().decode('utf-8')
        except UnicodeDecodeError:
            csv_file.seek(0)
            file_data = csv_file.read().decode('utf-8-sig')  # Handle BOM
        
        csv_reader = csv.DictReader(io.StringIO(file_data))
        
        # Validate required columns
        if 'question_text' not in csv_reader.fieldnames:
            errors.append('CSV file must contain "question_text" column')
            return imported, errors
        
        # Get next order
        max_order = Question.objects.filter(quiz=quiz).aggregate(
            max_order=Max('order')
        )['max_order'] or 0
        
        with transaction.atomic():
            for row_num, row in enumerate(csv_reader, start=2):
                try:
                    question_text = row.get('question_text', '').strip()
                    if not question_text:
                        errors.append(f'Row {row_num}: Missing question text')
                        continue
                    
                    question_type = row.get('question_type', 'multiple_choice').strip().lower()
                    # Validate question type
                    valid_types = ['multiple_choice', 'true_false', 'short_answer', 'fill_blank', 'essay']
                    if question_type not in valid_types:
                        errors.append(f'Row {row_num}: Invalid question_type "{question_type}". Using "multiple_choice"')
                        question_type = 'multiple_choice'
                    
                    marks = 1
                    if row.get('marks'):
                        try:
                            marks = int(row.get('marks'))
                            if marks < 1:
                                marks = 1
                        except ValueError:
                            marks = 1
                    
                    difficulty = row.get('difficulty', 'medium').strip().lower()
                    if difficulty not in ['easy', 'medium', 'hard']:
                        difficulty = 'medium'
                    
                    explanation = row.get('explanation', '').strip() or None
                    
                    max_order += 1
                    
                    # Create question
                    question = Question(
                        quiz=quiz,
                        question_type=question_type,
                        question_text=question_text,
                        marks=marks,
                        difficulty=difficulty,
                        order=max_order,
                        explanation=explanation,
                        school=school,
                    )
                    
                    # Handle question type specific data
                    if question_type in ['short_answer', 'fill_blank']:
                        correct_answer = row.get('correct_answer', '').strip()
                        if correct_answer:
                            question.correct_answer_text = correct_answer
                        question.save()
                    
                    elif question_type == 'essay':
                        # Essay questions don't need answer choices or correct answers
                        question.save()
                    
                    elif question_type in ['multiple_choice', 'true_false']:
                        question.save()
                        
                        # Get choices
                        choices = []
                        if question_type == 'true_false':
                            # True/False always has True and False
                            choices = ['True', 'False']
                        else:
                            # Multiple choice - get from choice1, choice2, etc.
                            for i in range(1, 11):  # Support up to 10 choices
                                choice_key = f'choice{i}'
                                if choice_key in csv_reader.fieldnames:
                                    choice_text = row.get(choice_key, '').strip()
                                    if choice_text:
                                        choices.append(choice_text)
                            
                            if len(choices) < 2:
                                errors.append(f'Row {row_num}: Multiple choice questions need at least 2 choices')
                                question.delete()
                                continue
                        
                        # Get correct answer
                        correct_answer = row.get('correct_answer', '').strip()
                        correct_indices = set()
                        
                        if question_type == 'true_false':
                            # For True/False, correct_answer can be 'true', 'false', '1', '2', 'True', 'False'
                            if correct_answer.lower() in ['true', 't', '1']:
                                correct_indices.add(0)
                            elif correct_answer.lower() in ['false', 'f', '2']:
                                correct_indices.add(1)
                            else:
                                # Default to first choice (True)
                                correct_indices.add(0)
                        else:
                            # For multiple choice, correct_answer can be:
                            # - Single index: "1", "2", etc.
                            # - Multiple indices: "1,2,3" or "1;2;3"
                            if correct_answer:
                                # Try to parse as comma or semicolon separated
                                for sep in [',', ';', '|']:
                                    if sep in correct_answer:
                                        parts = correct_answer.split(sep)
                                        for part in parts:
                                            part = part.strip()
                                            if part.isdigit():
                                                idx = int(part) - 1
                                                if 0 <= idx < len(choices):
                                                    correct_indices.add(idx)
                                        break
                                else:
                                    # Single value
                                    if correct_answer.isdigit():
                                        idx = int(correct_answer) - 1
                                        if 0 <= idx < len(choices):
                                            correct_indices.add(idx)
                                    else:
                                        # Try to match by text
                                        for idx, choice in enumerate(choices):
                                            if choice.strip().lower() == correct_answer.strip().lower():
                                                correct_indices.add(idx)
                                                break
                            
                            # If no correct answer specified, default to first
                            if not correct_indices:
                                correct_indices.add(0)
                        
                        # Create answer choices
                        for idx, choice_text in enumerate(choices):
                            AnswerChoice.objects.create(
                                question=question,
                                choice_text=choice_text,
                                is_correct=(idx in correct_indices),
                                order=idx + 1,
                                school=school,
                            )
                    
                    imported += 1
                
                except Exception as e:
                    errors.append(f'Row {row_num}: {str(e)}')
                    import traceback
                    errors.append(f'Row {row_num}: Traceback: {traceback.format_exc()}')
    
    except Exception as e:
        errors.append(f'Error reading CSV file: {str(e)}')
        import traceback
        errors.append(f'Traceback: {traceback.format_exc()}')
    
    return imported, errors


def _import_questions_from_excel(quiz, excel_file, school):
    """
    Helper function to import questions from Excel file.
    
    Uses the same format as CSV but reads from Excel.
    
    Returns: (imported_count, errors_list)
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required for Excel import. Install it with: pip install openpyxl")
    
    imported = 0
    errors = []
    
    try:
        excel_file.seek(0)
        wb = load_workbook(excel_file, read_only=True, data_only=True)
        ws = wb.active
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
            else:
                break
        
        if not headers or 'question_text' not in headers:
            errors.append('Excel file must contain "question_text" column in the first row')
            wb.close()
            return imported, errors
        
        # Get next order
        max_order = Question.objects.filter(quiz=quiz).aggregate(
            max_order=Max('order')
        )['max_order'] or 0
        
        # Convert Excel rows to dictionary format (same as CSV)
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell for cell in row if cell is not None):  # Skip empty rows
                row_dict = {}
                for idx, header in enumerate(headers):
                    if idx < len(row):
                        value = row[idx]
                        if value is not None:
                            row_dict[header] = str(value).strip()
                        else:
                            row_dict[header] = ''
                    else:
                        row_dict[header] = ''
                rows.append(row_dict)
        
        wb.close()
        
        # Process rows using the same logic as CSV
        with transaction.atomic():
            for row_num, row in enumerate(rows, start=2):
                try:
                    question_text = row.get('question_text', '').strip()
                    if not question_text:
                        errors.append(f'Row {row_num}: Missing question text')
                        continue
                    
                    question_type = row.get('question_type', 'multiple_choice').strip().lower()
                    valid_types = ['multiple_choice', 'true_false', 'short_answer', 'fill_blank', 'essay']
                    if question_type not in valid_types:
                        errors.append(f'Row {row_num}: Invalid question_type "{question_type}". Using "multiple_choice"')
                        question_type = 'multiple_choice'
                    
                    marks = 1
                    if row.get('marks'):
                        try:
                            marks = int(float(row.get('marks')))  # Excel numbers might be float
                            if marks < 1:
                                marks = 1
                        except (ValueError, TypeError):
                            marks = 1
                    
                    difficulty = row.get('difficulty', 'medium').strip().lower()
                    if difficulty not in ['easy', 'medium', 'hard']:
                        difficulty = 'medium'
                    
                    explanation = row.get('explanation', '').strip() or None
                    
                    max_order += 1
                    
                    # Create question
                    question = Question(
                        quiz=quiz,
                        question_type=question_type,
                        question_text=question_text,
                        marks=marks,
                        difficulty=difficulty,
                        order=max_order,
                        explanation=explanation,
                        school=school,
                    )
                    
                    # Handle question type specific data (same logic as CSV)
                    if question_type in ['short_answer', 'fill_blank']:
                        correct_answer = row.get('correct_answer', '').strip()
                        if correct_answer:
                            question.correct_answer_text = correct_answer
                        question.save()
                    
                    elif question_type == 'essay':
                        question.save()
                    
                    elif question_type in ['multiple_choice', 'true_false']:
                        question.save()
                        
                        # Get choices
                        choices = []
                        if question_type == 'true_false':
                            choices = ['True', 'False']
                        else:
                            for i in range(1, 11):
                                choice_key = f'choice{i}'
                                if choice_key in headers:
                                    choice_text = row.get(choice_key, '').strip()
                                    if choice_text:
                                        choices.append(choice_text)
                            
                            if len(choices) < 2:
                                errors.append(f'Row {row_num}: Multiple choice questions need at least 2 choices')
                                question.delete()
                                continue
                        
                        # Get correct answer
                        correct_answer = row.get('correct_answer', '').strip()
                        correct_indices = set()
                        
                        if question_type == 'true_false':
                            if correct_answer.lower() in ['true', 't', '1']:
                                correct_indices.add(0)
                            elif correct_answer.lower() in ['false', 'f', '2']:
                                correct_indices.add(1)
                            else:
                                correct_indices.add(0)
                        else:
                            if correct_answer:
                                for sep in [',', ';', '|']:
                                    if sep in correct_answer:
                                        parts = correct_answer.split(sep)
                                        for part in parts:
                                            part = part.strip()
                                            try:
                                                idx = int(float(part)) - 1  # Handle Excel numbers
                                                if 0 <= idx < len(choices):
                                                    correct_indices.add(idx)
                                            except (ValueError, TypeError):
                                                pass
                                        break
                                else:
                                    try:
                                        idx = int(float(correct_answer)) - 1
                                        if 0 <= idx < len(choices):
                                            correct_indices.add(idx)
                                    except (ValueError, TypeError):
                                        for idx, choice in enumerate(choices):
                                            if choice.strip().lower() == correct_answer.strip().lower():
                                                correct_indices.add(idx)
                                                break
                            
                            if not correct_indices:
                                correct_indices.add(0)
                        
                        # Create answer choices
                        for idx, choice_text in enumerate(choices):
                            AnswerChoice.objects.create(
                                question=question,
                                choice_text=choice_text,
                                is_correct=(idx in correct_indices),
                                order=idx + 1,
                                school=school,
                            )
                    
                    imported += 1
                
                except Exception as e:
                    errors.append(f'Row {row_num}: {str(e)}')
                    import traceback
                    errors.append(f'Row {row_num}: Traceback: {traceback.format_exc()}')
    
    except Exception as e:
        errors.append(f'Error reading Excel file: {str(e)}')
        import traceback
        errors.append(f'Traceback: {traceback.format_exc()}')
    
    return imported, errors


@login_required
@require_http_methods(["GET"])
def question_import_template_view(request, quiz_id, format_type='csv'):
    """
    Download a template file for question import.
    
    Args:
        format_type: 'csv' or 'xlsx'
    """
    if request.user.role not in ["teacher", "admin"]:
        messages.error(request, "You don't have permission to perform this action.")
        return redirect("quiz_app:quiz_list")
    
    if not request.user.teacher_profile:
        messages.error(request, "Teacher profile not found. Please contact administrator.")
        return redirect("quiz_app:quiz_list")
    
    teacher = request.user.teacher_profile
    school = request.user.school
    quiz = get_object_or_404(Quiz, pk=quiz_id, teacher=teacher, school=school)
    
    if format_type == 'csv':
        # Create CSV template
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="question_import_template.csv"'
        
        # Write BOM for Excel compatibility
        response.write('\ufeff')
        
        writer = csv.writer(response)
        
        # Write headers
        headers = [
            'question_text',
            'question_type',
            'marks',
            'difficulty',
            'correct_answer',
            'choice1',
            'choice2',
            'choice3',
            'choice4',
            'choice5',
            'explanation'
        ]
        writer.writerow(headers)
        
        # Write example rows
        examples = [
            [
                'What is the capital of France?',
                'multiple_choice',
                '1',
                'easy',
                '3',
                'London',
                'Berlin',
                'Paris',
                'Madrid',
                '',
                'Paris is the capital and largest city of France.'
            ],
            [
                'Python is a programming language.',
                'true_false',
                '1',
                'easy',
                'true',
                '',
                '',
                '',
                '',
                '',
                'Python is indeed a programming language.'
            ],
            [
                'What is 2 + 2?',
                'short_answer',
                '1',
                'easy',
                '4',
                '',
                '',
                '',
                '',
                '',
                'The answer is 4.'
            ],
            [
                'The capital of England is _____.',
                'fill_blank',
                '1',
                'easy',
                'London',
                '',
                '',
                '',
                '',
                '',
                'London is the capital of England.'
            ]
        ]
        
        for example in examples:
            writer.writerow(example)
        
        return response
    
    elif format_type == 'xlsx':
        # Create Excel template
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messages.error(request, 'Excel template requires openpyxl. Please install it or use CSV format.')
            return redirect("quiz_app:quiz_detail", quiz_id=quiz_id)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Question Import Template"
        
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = [
            'question_text',
            'question_type',
            'marks',
            'difficulty',
            'correct_answer',
            'choice1',
            'choice2',
            'choice3',
            'choice4',
            'choice5',
            'explanation'
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Example rows
        examples = [
            [
                'What is the capital of France?',
                'multiple_choice',
                1,
                'easy',
                3,
                'London',
                'Berlin',
                'Paris',
                'Madrid',
                '',
                'Paris is the capital and largest city of France.'
            ],
            [
                'Python is a programming language.',
                'true_false',
                1,
                'easy',
                'true',
                '',
                '',
                '',
                '',
                '',
                'Python is indeed a programming language.'
            ],
            [
                'What is 2 + 2?',
                'short_answer',
                1,
                'easy',
                '4',
                '',
                '',
                '',
                '',
                '',
                'The answer is 4.'
            ],
            [
                'The capital of England is _____.',
                'fill_blank',
                1,
                'easy',
                'London',
                '',
                '',
                '',
                '',
                '',
                'London is the capital of England.'
            ]
        ]
        
        for row_idx, example in enumerate(examples, start=2):
            for col_idx, value in enumerate(example, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="question_import_template.xlsx"'
        wb.save(response)
        return response
    
    else:
        messages.error(request, 'Invalid format type. Please use "csv" or "xlsx".')
        return redirect("quiz_app:quiz_detail", quiz_id=quiz_id)


@login_required
@require_http_methods(["GET", "POST"])
def question_bulk_delete_view(request, quiz_id):
    """
    Bulk delete questions.
    
    GET: Returns bulk delete modal
    POST: Deletes multiple questions by IDs
    """
    if request.user.role not in ["teacher", "admin"]:
        if request.method == "GET":
            return JsonResponse({'error': 'Permission denied'}, status=403)
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if not request.user.teacher_profile:
        if request.method == "GET":
            return JsonResponse({'error': 'Teacher profile not found'}, status=404)
        return JsonResponse({'error': 'Teacher profile not found'}, status=404)
    
    teacher = request.user.teacher_profile
    school = request.user.school
    quiz = get_object_or_404(Quiz, pk=quiz_id, teacher=teacher, school=school)
    
    if request.method == "GET":
        # Check if this is an AJAX request for modal
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            html = render(request, 'quiz/question/partials/bulk_delete_modal.html', {
                'quiz': quiz,
            }).content.decode('utf-8')
            return JsonResponse({'html': html})
        
        # Regular page request
        return redirect("quiz_app:quiz_detail", quiz_id=quiz_id)
    
    # POST - Bulk delete
    try:
        data = json.loads(request.body)
        question_ids = data.get('question_ids', [])
        
        if not question_ids:
            return JsonResponse({
                'success': False,
                'error': 'No questions selected.'
            }, status=400)
        
        # Convert to integers and filter out invalid IDs
        try:
            question_ids = [int(qid) for qid in question_ids if qid]
        except (ValueError, TypeError):
            return JsonResponse({
                'success': False,
                'error': 'Invalid question IDs provided.'
            }, status=400)
        
        questions = Question.objects.filter(
            pk__in=question_ids,
            quiz=quiz,
            school=school
        )
        
        count = questions.count()
        if count == 0:
            return JsonResponse({
                'success': False,
                'error': 'No valid questions found to delete.'
            }, status=400)
        
        with transaction.atomic():
            # Delete questions (this will cascade delete answer choices)
            questions.delete()
            
            # Recalculate quiz total marks
            quiz.calculate_total_marks()
        
        return JsonResponse({
            'success': True,
            'message': f'{count} question(s) deleted successfully.'
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error deleting questions: {str(e)}'
        }, status=500)

